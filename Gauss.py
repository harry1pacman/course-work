#имя проекта: gauss
#номер версии: 1.0
#имя файла: Gauss.py
#автор и его учебная группа: И. Поллак, ЭУ-120
#дата создания: 24.05.2020
# дата последней модификации: 24.05.2020
#связанные файлы: numpy, openpyxl
# описание: решение СЛАУ с использованием библиотеки numpy и openpyxl
#версия Python: 3.8
import numpy as np
import openpyxl


def FancyPrint(A, B, selected):
    for row in range(len(B)):
        print("(", end='')
        for col in range(len(A[row])):
            print("\t{1:10.2f}{0}".format(" " if (selected is None or selected != (row, col)) else "*", A[row][col]),
                  end='')
        print("\t) * (\tX{0}) = (\t{1:10.2f})".format(row + 1, B[row]))


wb = openpyxl.load_workbook("D:\gauss.xlsx")
sheet = wb.get_sheet_by_name('Лист1')

f = open("D:\gauss.txt", 'w')
row_ex = [1, 8, 15, 19, 22]

for row in row_ex:
    if row == 1 or 8:
        myA = np.empty([6, 6], dtype=float)

        for i in range(0, 6):
            for j in range(0, 6):
                myA.itemset((i, j), sheet.cell(row=(row + i), column=(j + 1)).value)

        myB = np.empty(6, dtype=float)

        for i in range(0, 6):
            myB.itemset(i, sheet.cell(row=(row + i), column=7).value)


    if row == 15:
        myA = np.empty([3, 3], dtype=float)

        for i in range(0, 3):
            for j in range(0, 3):
                myA.itemset((i, j), sheet.cell(row=(row + i), column=(j + 1)).value)

        myB = np.empty(3, dtype=float)

        for i in range(0, 3):
            myB.itemset(i, sheet.cell(row=(row + i), column=4).value)


    if row == 19:
        myA = np.empty([2, 2], dtype=float)

        for i in range(0, 2):
            for j in range(0, 2):
                myA.itemset((i, j), sheet.cell(row=(row + i), column=(j + 1)).value)

        myB = np.empty(2, dtype=float)

        for i in range(0, 2):
            myB.itemset(i, sheet.cell(row=(row + i), column=3).value)


    if row == 22:
        myA = np.empty([5, 5], dtype=float)

        for i in range(0, 5):
            for j in range(0, 5):
                myA.itemset((i, j), sheet.cell(row=(row + i), column=(j + 1)).value)

        myB = np.empty(5, dtype=float)

        for i in range(0, 5):
            myB.itemset(i, sheet.cell(row=(row + i), column=6).value)


    print("Исходная система:")
    FancyPrint(myA, myB, float)
    slv = np.linalg.solve(myA, myB)

    print("Решаем:")
    print(slv)
    f.write(str(slv) + '\n')
    print("\n\n")

f.close()
wb.close()
